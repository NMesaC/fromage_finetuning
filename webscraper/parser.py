# Imports
import openpyxl
import json
import requests
import PIL.Image

from openpyxl.worksheet.worksheet import Worksheet

from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def main():
    # Open the file and load its contents into a dictionary
    with open("scizor_res.json", 'r') as file:
        data_dict = json.load(file)

    # Call a Workbook() function of openpyxl to create a new blank Workbook object
    wb = openpyxl.Workbook()

    # Get workbook active sheet from the active attribute
    sheet = wb.active
    assert isinstance(sheet, Worksheet)

    # Set new sizes for the columns
    col_size, row_size = 50, 100
    sheet.column_dimensions['A'].width = col_size
    sheet.column_dimensions['B'].width = col_size

    # Set new sizes for the rows
    for i in range(1, 101):
        sheet.row_dimensions[i].height = row_size

    # Create a for loop so each row of the file of the format:
    # description | url | image
    # This is just for labelling purposes, the actual file should be:
    # description | url
    row = 1
    col = 1
    print("Parsing json data...")
    for res in data_dict:
        # Description Cell
        col = 1
        c1  = sheet.cell(row = row, column = col)
        c1.value = "temp_disc"

        # Image Cell
        col = 3
        url = res['original']
        response = requests.get(url)
        if response.status_code != 200:
            print(f"Failed to download image. Status code: {response.status_code}")
            print(f"Url = {url}")
            continue
        try:
            img_data = BytesIO(response.content)
            pil_image = PIL.Image.open(img_data)
            if pil_image.mode == 'RGBA':
                pil_image = pil_image.convert('RGB')
            img_byte_arr = BytesIO()
            pil_image.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            img = Image(img_byte_arr)
        except:
            print(f"Error loading image on row {row}")
            continue
        width, height = pil_image.size
        max_width, max_height = 2*col_size, 2*row_size
        scale_factor = min(max_width/width, max_height/height)
        img.width = int(width * scale_factor)
        img.height = int(height * scale_factor)
        cell_address = f"{get_column_letter(col)}{row}"
        sheet.add_image(img, cell_address)

        # URL Cell
        col = 2
        c3 = sheet.cell(row = row, column = col)
        c3.value = url

        #Increment row
        row += 1
    wb.save("scizor.xlsx")
    print("Done!")



if __name__ == "__main__":
    main()
